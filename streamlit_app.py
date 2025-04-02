import pandas as pd
import streamlit as st
import numpy as np
import io
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pybaseball import statcast_pitcher, statcast_batter, playerid_lookup
from datetime import datetime
from pytz import timezone
from functools import lru_cache

st.set_page_config(page_title="Pitcher vs Batter Analyzer", layout="wide")
st.title("⚾ Pitcher vs Batter Matchup Analyzer (Statcast Live)")

st.markdown("""
Enter the names of a pitcher and a batter to pull real-time Statcast data and generate a pitch-by-pitch matchup breakdown:
- Strikeout probability metrics: **K%**, **PutAway%**, **Whiff%**
- Contact effectiveness: **OBA**, **BA**, **SLG**
- Advantage deltas highlighted in **red (pitcher)** or **green (batter)**
""")

# --- Input section ---
pitcher_name = st.text_input("Pitcher Name (First Last)", "Corbin Burnes")
batter_name = st.text_input("Batter Name (First Last)", "Jazz Chisholm")

# Default end date = today in Eastern Time
et_now = datetime.now(timezone("US/Eastern"))
start_date = st.date_input("Start Date", value=datetime(2024, 4, 1))
end_date = st.date_input("End Date", value=et_now.date())

# --- Utility functions ---
def get_player_id(name):
    try:
        first, last = name.strip().split()
        result = playerid_lookup(last, first)
        if not result.empty:
            return int(result.iloc[0]['key_mlbam'])
    except:
        return None
    return None

@lru_cache(maxsize=16)
def get_pitcher_data(pid, start, end):
    return statcast_pitcher(start, end, pid)

@lru_cache(maxsize=16)
def get_batter_data(bid, start, end):
    return statcast_batter(start, end, bid)

# --- Stat calculation helpers ---
def calculate_pitch_stats(df, role):
    grouped = df.groupby("pitch_type")
    stats = grouped.agg({
        "description": lambda x: sum(x.str.contains("strikeout", case=False)),
        "events": lambda x: sum(x.dropna().isin(["strikeout"])),
        "pitch_type": "count",
        "release_speed": "mean",
        "balls": "mean",
        "strikes": "mean",
        "estimated_ba_using_speedangle": "mean",
        "estimated_woba_using_speedangle": "mean",
        "estimated_slg_using_speedangle": "mean",
        "description": lambda x: sum(x.str.contains("swinging_strike", case=False)),
    }).rename(columns={
        "description": "Whiffs",
        "events": "Ks",
        "pitch_type": "Pitches"
    })
    stats["K%"] = stats["Ks"] / stats["Pitches"] * 100
    stats["Whiff%"] = stats["Whiffs"] / stats["Pitches"] * 100
    stats["PutAway%"] = stats["Ks"] / stats["Pitches"] * 100
    stats["OBA"] = stats["estimated_woba_using_speedangle"]
    stats["BA"] = stats["estimated_ba_using_speedangle"]
    stats["SLG"] = stats["estimated_slg_using_speedangle"]
    stats = stats[["K%", "Whiff%", "PutAway%", "OBA", "BA", "SLG"]]
    stats.columns = [f"{col}_{role}" for col in stats.columns]
    return stats

# --- Color logic ---
def highlight_deltas(val):
    if pd.isnull(val):
        return "background-color: #333333; color: #ffffff;"

    try:
        val = float(val)
    except ValueError:
        return ""

    cap = 10
    intensity = min(abs(val), cap) / cap

    if val > 0:
        red = int(34 - 20 * intensity)
        green = int(85 + (170 * intensity))
        blue = int(34 - 20 * intensity)
    elif val < 0:
        red = int(102 + (153 * intensity))
        green = int(17 - 15 * intensity)
        blue = int(17 - 15 * intensity)
    else:
        return "background-color: #333333; color: #ffffff; font-weight: bold;"

    return f"background-color: rgb({red},{green},{blue}); color: #ffffff; font-weight: bold;"

# --- App logic ---
if st.button("Run Matchup Analysis"):
    pid = get_player_id(pitcher_name)
    bid = get_player_id(batter_name)

    if not pid or not bid:
        st.error("Could not find one or both players. Please check the names.")
    else:
        with st.spinner("Fetching and analyzing data..."):
            p_df = get_pitcher_data(pid, str(start_date), str(end_date))
            b_df = get_batter_data(bid, str(start_date), str(end_date))

            pitcher_stats = calculate_pitch_stats(p_df, "Pitcher")
            batter_stats = calculate_pitch_stats(b_df, "Batter")

            combined = pd.merge(pitcher_stats, batter_stats, left_index=True, right_index=True, how="inner")

            for stat in ["K%", "Whiff%", "PutAway%", "OBA", "BA", "SLG"]:
                combined[f"{stat} Delta"] = combined[f"{stat}_Pitcher"] - combined[f"{stat}_Batter"]

            pitch_name_map = {
                "CH": "Changeup",
                "CU": "Curveball",
                "FC": "Cutter",
                "FF": "Four-Seam Fastball",
                "SI": "Sinker",
                "SL": "Slider",
                "ST": "Sweeper"
            }
            combined.index = combined.index.map(lambda x: pitch_name_map.get(x, x))

            delta_cols = [col for col in combined.columns if "Delta" in col]
            styled = combined.style.applymap(highlight_deltas, subset=delta_cols).format("{:.2f}")

            st.subheader("📊 Matchup Table")
            st.dataframe(styled, use_container_width=True)

            st.download_button("Download CSV", combined.to_csv().encode("utf-8"), "matchup_analysis.csv")

            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Matchup"
            for r in dataframe_to_rows(combined.reset_index(), index=False, header=True):
                ws.append(r)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            wb.save(output)
            st.download_button("Download Excel", output.getvalue(), "matchup_analysis.xlsx")
